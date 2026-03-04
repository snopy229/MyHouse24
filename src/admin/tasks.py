from copy import copy
from datetime import datetime
from io import BytesIO

import openpyxl
from celery import shared_task
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.core.mail import send_mail, EmailMessage
from django.db.models import Sum, Q
from django.utils import timezone
from openpyxl.reader.excel import load_workbook
from weasyprint import HTML

from .models import CashBox, Receipt, XlsTemplate, ServiceFullCost


@shared_task
def send_password(raw_password, user_email):
    try:
        send_mail(
            subject="Создания пароля",
            message=f"Новый пароль: {raw_password}",
            from_email=None,
            recipient_list=[user_email],
        )
    except Exception as e:
        print(f"Ошибка отправки: {e}")


@shared_task
def send_invite(user_email):
    try:
        send_mail(
            subject="Тест",
            message=f"Пользователь:{user_email}, вернись на сайт",
            from_email=None,
            recipient_list=[user_email],
        )
    except Exception as e:
        print(f"Ошибка отправки: {e}")


@shared_task
def send_invite_owner(user_email):
    try:
        send_mail(
            subject="Приглошение",
            message="Заходите на наш сайт.",
            from_email=None,
            recipient_list=[user_email],
        )
    except Exception as e:
        print(f"Ошибка отправки: {e}")


@shared_task
def send_pdf_task(template_id, pk):
    receipt = Receipt.objects.get(pk=pk)
    wb = make_receipt_xls(template_id, pk)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    while max_row > 1:
        if any(ws.cell(max_row, c).value is not None for c in range(1, max_col + 1)):
            break
        max_row -= 1

    while max_col > 1:
        if any(ws.cell(r, max_col).value is not None for r in range(1, max_row + 1)):
            break
        max_col -= 1

    merged = {}
    for merge_range in ws.merged_cells.ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        max_r, max_c = merge_range.max_row, merge_range.max_col
        merged[(min_row, min_col)] = (max_r - min_row + 1, max_c - min_col + 1)
        for r in range(min_row, max_r + 1):
            for c in range(min_col, max_c + 1):
                if (r, c) != (min_row, min_col):
                    merged[(r, c)] = None

    col_widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        col_widths[col_idx] = int((col_dim.width or 8) * 7)

    html_rows = []
    for row_idx in range(1, max_row + 1):
        row_height = ws.row_dimensions[row_idx].height or 15
        cells_html = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            coord = (cell.row, cell.column)
            if coord in merged and merged[coord] is None:
                continue

            td_attrs = []
            style_parts = []

            if coord in merged:
                rowspan, colspan = merged[coord]
                if rowspan > 1:
                    td_attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    td_attrs.append(f'colspan="{colspan}"')

            width = col_widths.get(col_idx, 56)
            style_parts.append(f"width:{width}px")
            style_parts.append(f"height:{row_height}pt")

            if cell.font:
                if cell.font.bold:
                    style_parts.append("font-weight:bold")
                if cell.font.italic:
                    style_parts.append("font-style:italic")
                if cell.font.size:
                    style_parts.append(f"font-size:{cell.font.size}pt")
                if cell.font.color and cell.font.color.type == "rgb":
                    style_parts.append(f"color:#{cell.font.color.rgb[2:]}")

            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
                rgb = cell.fill.fgColor.rgb[2:]
                if rgb != "000000":
                    style_parts.append(f"background-color:#{rgb}")

            if cell.alignment:
                if cell.alignment.horizontal:
                    style_parts.append(f"text-align:{cell.alignment.horizontal}")
                if cell.alignment.vertical:
                    style_parts.append(f"vertical-align:{cell.alignment.vertical}")
                if cell.alignment.wrap_text:
                    style_parts.append("white-space:pre-wrap")

            value = cell.value
            if value is not None:
                if isinstance(value, datetime):
                    value = value.strftime("%d.%m.%Y")
                elif isinstance(value, float):
                    value = f"{value:,.2f}".replace(",", " ")

            td_style = "; ".join(style_parts)
            td_open = f'<td {" ".join(td_attrs)} style="{td_style}">'
            cells_html.append(f'{td_open}{value if value is not None else ""}</td>')

        html_rows.append(f'<tr>{"".join(cells_html)}</tr>')

    html_content = f"""
    <html><head><style>
        @page {{ size: A4 landscape; margin: 1cm; }}
        table {{ border-collapse: collapse; font-family: Arial, sans-serif; font-size: 10pt; }}
        td {{ border: 1px solid #ccc; padding: 3px; overflow: hidden; }}
    </style></head>
    <body><table>{"".join(html_rows)}</table></body></html>
    """

    pdf_buffer = BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    email = EmailMessage(
        subject="Ваш чек",
        body="Здравствуйте! Во вложении находится ваш чек в формате PDF.",
        from_email=None,
        to=[receipt.apartment.owner.email],
    )
    email.attach(
        filename=f"receipt_{pk}.pdf",
        content=pdf_buffer.getvalue(),
        mimetype="application/pdf",
    )
    email.send()


@shared_task
def generate_excel_task(template_id, pk):
    receipt = Receipt.objects.get(pk=pk)
    wb = make_receipt_xls(template_id, pk)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    now = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"receipt_{receipt.number}_{now}.xlsx"
    actual_filename = default_storage.save(filename, ContentFile(buffer.read()))
    file_url = default_storage.url(actual_filename)

    return file_url


def make_receipt_xls(template_id, pk):
    receipt = (
        Receipt.objects.select_related("bankbook", "apartment__owner", "house")
        .prefetch_related("servicefullcost_set")
        .get(pk=pk)
    )
    template = XlsTemplate.objects.get(pk=template_id)
    stats = CashBox.objects.filter(bank_book=receipt.bankbook).aggregate(
        total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
        total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
    )
    receipt_cost = sum(a.full_cost for a in receipt.servicefullcost_set.all())

    receipts = Receipt.objects.filter(bankbook=receipt.bankbook)

    receipts_pay = (
        ServiceFullCost.objects.filter(receipt__in=receipts).aggregate(
            total=Sum("full_cost")
        )["total"]
        or 0
    )
    balance = (stats["total_cm"] or 0) - receipts_pay - (stats["total_og"] or 0)
    cost = receipt_cost - balance
    wb = load_workbook(template.template.path)
    ws = wb.active
    data_map = {
        "{bb_number}": receipt.bankbook.number if receipt.bankbook else "",
        "{rp_number}": receipt.number,
        "{rp_date}": receipt.date.strftime("%d.%m.%Y"),
        "{owner_flat}": f'{receipt.apartment.owner.fullname if receipt.apartment.owner else ''} {receipt.house.title if receipt.house else ''}, {receipt.apartment.number}',
        "{rp_cost}": sum(a.full_cost for a in receipt.servicefullcost_set.all()),
        "{bb_balance}": balance,
        "{cost}": f'{cost if cost > 0 else ''}',
        "{date_now}": timezone.now().date().strftime("%d.%m.%Y"),
        "{date_interval}": f'{receipt.date_from.strftime('%d.%m.%Y')}-{receipt.date_to.strftime('%d.%m.%Y')}',
    }
    services = ServiceFullCost.objects.filter(receipt=receipt)
    fill_service_table(ws, services, receipt)
    fill_template(ws, data_map)

    return wb


COLUMN_MAP = {
    "service__title": ("service_title", "Услуга"),
    "tariff": ("receipt", "Тариф"),
    "unit": ("service_unit", "Ед.изм"),
    "consumption": ("service", "Расход"),
    "full_cost": ("service", "Сумма, грн"),
}


def get_column_indices(ws, mapping):
    indices = {}
    header_row_idx = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if not cell.value:
                continue

            for field, (source, header_text) in mapping.items():
                if header_text in str(cell.value):
                    indices[field] = cell.column
                    header_row_idx = cell.row

        if len(indices) == len(mapping):
            break

    return indices, header_row_idx


def fill_service_table(ws, services, receipt):
    services = list(services)

    # Найти строку с плейсхолдерами сервиса
    template_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "{service}" in str(cell.value):
                template_row = cell.row
                break
        if template_row:
            break

    if not template_row:
        return None

    placeholder_cols = {}
    for cell in ws[template_row]:
        if cell.value:
            placeholder_cols[str(cell.value).strip()] = cell.column

    ws.delete_rows(template_row, 1)
    ws.insert_rows(template_row, amount=len(services))
    template_styles = {}

    for cell in ws[template_row]:
        template_styles[cell.column] = {
            "font": copy(cell.font),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "fill": copy(cell.fill),
            "number_format": cell.number_format,
        }

    ws.delete_rows(template_row, 1)
    ws.insert_rows(template_row, amount=len(services))

    for i, service in enumerate(services):
        row_idx = template_row + i

        mapping = {
            "{service}": service.service.title if service.service else "",
            "{tariff}": receipt.tariff.title if receipt.tariff else "",
            "{unit_of_measure}": service.unit.units if service.service else "",
            "{consumption}": getattr(service, "consumption", ""),
            "{fullcost}": getattr(service, "full_cost", ""),
        }

        for placeholder, value in mapping.items():
            col = placeholder_cols.get(placeholder)
            if col:
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value
                style = template_styles.get(col)
                if style:
                    cell.font = style["font"]
                    cell.border = style["border"]
                    cell.alignment = style["alignment"]
                    cell.fill = style["fill"]
                    cell.number_format = style["number_format"]

    return template_row + len(services)


def fill_template(ws, data_map):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, value in data_map.items():
                    if key in cell.value:
                        cell.value = cell.value.replace(key, str(value))
