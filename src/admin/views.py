import json

from ajax_datatable import AjaxDatatableView
from celery.result import AsyncResult
from django.contrib.auth.mixins import PermissionRequiredMixin, AccessMixin
from django.db import transaction, IntegrityError
from django.db.models import CharField, Value, Sum, Q, FloatField, F
from django.db.models.functions import Concat, Coalesce, TruncMonth
from django.forms import inlineformset_factory
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.views import View
from django.views.generic import (
    CreateView,
    UpdateView,
    TemplateView,
    DeleteView,
    DetailView,
    ListView,
    FormView,
)
from django.views.generic.edit import FormMixin
from openpyxl import Workbook
from datetime import datetime
from django.utils import timezone

from src.settings.models import Tariffs, Service, ServicesCost
from src.user.models import User, Role
from src.user.choices import Status
from src.user.utils import send_email_verify
from .choices import (
    StatusBankBook,
    StatusCounter,
    StatusCall,
    StatusReceipt,
    StatusCashBox,
)
from .forms import (
    FloorFormSet,
    StaffFormSet,
    ApartmentForm,
    HouseForm,
    SectionFormSet,
    OwnerForm,
    BankBookForm,
    CounterForm,
    MasterCallForm,
    ReceiptForm,
    ServiceFullCostFormSet,
    MessageForm,
    CashBoxForm,
    InviteForm,
    ServiceFullCostForm,
    XlsTemplateForm,
    XlsTemplateShowForm,
)
from src.admin.models import (
    House,
    Apartment,
    BankBook,
    Counter,
    MasterCall,
    Receipt,
    Message,
    MessageStatus,
    CashBox,
    ServiceFullCost,
    XlsTemplate,
)
from .tasks import send_password, send_invite_owner, generate_excel_task, send_pdf_task


class RedirectMixin(AccessMixin):
    redirect_url = reverse_lazy("admin:error")

    def handle_no_permission(self):
        if self.permission_denied_message:
            pass
        return redirect(self.redirect_url)


class CommonContextMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        stats = CashBox.objects.aggregate(
            total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
            total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
        )
        balance = (stats["total_cm"] or 0) - (stats["total_og"] or 0)

        formatted_balance = "{:,.2f}".format(balance).replace(",", " ")

        context["cashbox_status"] = formatted_balance
        stats = CashBox.objects.filter(bank_book__isnull=False).aggregate(
            total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
            total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
        )

        receipts = Receipt.objects.filter(bankbook__isnull=False)

        receipts_pay = (
            ServiceFullCost.objects.filter(receipt__in=receipts).aggregate(
                total=Sum("full_cost")
            )["total"]
            or 0
        )
        balance = (stats["total_cm"] or 0) - receipts_pay - (stats["total_og"] or 0)
        formatted_balance = "{:,.2f}".format(balance).replace(",", " ")
        context["bankbook_status"] = formatted_balance
        bankbooks = BankBook.objects.all()
        dept = 0
        for bankbook in bankbooks:
            stats = CashBox.objects.filter(bank_book=bankbook).aggregate(
                total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
                total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
            )

            receipts = Receipt.objects.filter(bankbook=bankbook)

            receipts_pay = (
                ServiceFullCost.objects.filter(receipt__in=receipts).aggregate(
                    total=Sum("full_cost")
                )["total"]
                or 0
            )
            balance = (stats["total_cm"] or 0) - receipts_pay - (stats["total_og"] or 0)
            if balance < 0:
                balance = -balance
                dept += balance

        formatted_dept = "{:,.2f}".format(dept).replace(",", " ")
        context["bankbook_dept"] = formatted_dept

        return context


class ErrorPage(TemplateView):
    template_name = "admin_error.html"


class Statistic(
    RedirectMixin, CommonContextMixin, PermissionRequiredMixin, TemplateView
):
    template_name = "statistic.html"
    permission_required = "role.has_statistics"

    def get_context_data(self, request=None, **kwargs):
        context = super().get_context_data()
        houses = House.objects.all()
        excluded_roles = [
            "Директор",
        ]
        if request.user.role.title not in excluded_roles:
            houses = houses.filter(owner=request.user)
        context["houses"] = houses.count()
        owners = User.objects.filter(status="active", is_staff=False)
        if request.user.role.title not in excluded_roles:
            owners = owners.filter(apartment_set__house__owner=request.user)
        context["owners"] = owners.count()
        master_call_in_work = MasterCall.objects.filter(call_status="IN WORK")
        if request.user.role.title not in excluded_roles:
            master_call_in_work = master_call_in_work.filter(master=request.user)
        context["master_calls_in_work"] = master_call_in_work.count()
        apartments = Apartment.objects.all()
        if request.user.role.title not in excluded_roles:
            apartments = apartments.filter(house__owner=request.user)
        context["apartments"] = apartments.count()
        bankbook = BankBook.objects.all()
        if request.user.role.title not in excluded_roles:
            bankbook = bankbook.filter(house__owner=request.user)
        context["bankbooks"] = bankbook.count()
        new_master_call = MasterCall.objects.filter(call_status="NEW")
        if request.user.role.title not in excluded_roles:
            new_master_call = new_master_call.filter(master=request.user)
        context["new_master_call"] = new_master_call.count()
        monthly_stats = (
            ServiceFullCost.objects.all()
            .annotate(month=TruncMonth("receipt__date"))
            .values("month")
            .annotate(total=Sum("full_cost"))
            .order_by("month")
        )
        coming_money = (
            CashBox.objects.filter(status="CM", receipt__isnull=False)
            .annotate(month=TruncMonth("receipt__date"))
            .values("month")
            .annotate(total=Sum("sum"))
            .order_by("month")
        )
        bar_data = [item["total"] for item in monthly_stats]
        income_data = [item["total"] for item in coming_money]
        context["bar_data"] = bar_data
        context["income_data"] = income_data
        cb_coming = (
            CashBox.objects.filter(status="CM")
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total=Sum("sum"))
            .order_by("month")
        )
        cb_outgo = (
            CashBox.objects.filter(status="OG")
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total=Coalesce(Sum("sum"), 0.0))
            .order_by("month")
        )
        cm_list = [item["total"] for item in cb_coming]
        og_list = [item["total"] for item in cb_outgo]
        context["cb_coming"] = cm_list
        context["cb_outgo"] = og_list
        return context


class CreateHouse(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = House
    form_class = HouseForm
    context_object_name = "house"
    template_name = "house.html"
    permission_required = "role.has_houses"
    success_url = reverse_lazy("admin:house-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["sections"] = SectionFormSet(self.request.POST, prefix="sections")
            context["floors"] = FloorFormSet(self.request.POST, prefix="floors")
            context["staff"] = StaffFormSet(self.request.POST, prefix="staff")
        else:
            context["sections"] = SectionFormSet(prefix="sections")
            context["floors"] = FloorFormSet(prefix="floors")
            context["staff"] = StaffFormSet(prefix="staff")
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        sections = context["sections"]
        floors = context["floors"]
        staff = context["staff"]

        if sections.is_valid() and floors.is_valid() and staff.is_valid():
            with transaction.atomic():
                self.object = form.save()
                sections.instance = self.object
                floors.instance = self.object
                staff.instance = self.object
                sections.save()
                floors.save()
                staff.save()
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class EditHouse(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = House
    form_class = HouseForm
    context_object_name = "house"
    template_name = "house.html"
    permission_required = "role.has_houses"
    success_url = reverse_lazy("admin:house-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["sections"] = SectionFormSet(
                self.request.POST, instance=self.object, prefix="sections"
            )
            context["floors"] = FloorFormSet(
                self.request.POST, instance=self.object, prefix="floors"
            )
            context["staff"] = StaffFormSet(
                self.request.POST, instance=self.object, prefix="staff"
            )
        else:
            context["sections"] = SectionFormSet(
                instance=self.object, prefix="sections"
            )
            context["floors"] = FloorFormSet(instance=self.object, prefix="floors")
            context["staff"] = StaffFormSet(instance=self.object, prefix="staff")
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        sections = context["sections"]
        floors = context["floors"]
        staff = context["staff"]

        if sections.is_valid() and floors.is_valid() and staff.is_valid():
            with transaction.atomic():
                self.object = form.save()
                sections.instance = self.object
                floors.instance = self.object
                staff.instance = self.object
                sections.save()
                floors.save()
                staff.save()
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class DetailHouse(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = House
    template_name = "house_detail.html"
    context_object_name = "house"
    permission_required = "role.has_houses"


class HouseAjaxTable(AjaxDatatableView):
    model = House
    title = "Дома"
    initial_order = [["id", "asc"]]
    column_defs = [
        {"name": "id", "visible": True, "searchable": False, "title": "#"},
        {
            "name": "title",
            "visible": True,
            "searchable": True,
            "title": "Название",
        },
        {
            "name": "address",
            "visible": True,
            "searchable": True,
            "title": "Адрес",
        },
        {
            "name": "actions",
            "visible": True,
            "searchable": False,
            "orderable": False,
            "title": "",
        },
    ]

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset()
        excluded_roles = [
            "Директор",
        ]
        if request.user.role.title not in excluded_roles:
            qs = qs.filter(owner=request.user)
        return qs

    def customize_row(self, row, obj):
        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="/admin/house/edit/{0}/" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="/admin/house/delete/{0}/" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            obj.id,
        )

        detail_url = reverse("admin:detail-house", args=[obj.id])

        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}
        return


class DeleteHouse(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = House
    permission_required = "role.has_houses"

    def get(self, request, pk):
        article = get_object_or_404(House, pk=pk)
        article.delete()
        return redirect("admin:house-list")


class HouseList(RedirectMixin, PermissionRequiredMixin, ListView):
    template_name = "houses.html"
    model = House
    context_object_name = "house"
    permission_required = "role.has_houses"


class CreateFlat(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = Apartment
    template_name = "flat.html"
    form_class = ApartmentForm
    permission_required = "role.has_apartments"

    def form_valid(self, form):
        self.object = form.save()

        if "save_and_copy" in self.request.POST:
            saved_data = {
                "house": self.object.house.id,
                "section": self.object.section.id,
                "floor": self.object.floor.id,
                "tariff": self.object.tariff.id,
            }
            self.request.session["copied_data"] = saved_data

            return redirect("admin:create-flat")

        return redirect("admin:flat-list")

    def get_initial(self):
        initial = super().get_initial()

        copied_data = self.request.session.pop("copied_data", None)
        if copied_data:
            initial.update(copied_data)

        return initial


class ListFlat(RedirectMixin, PermissionRequiredMixin, TemplateView):
    template_name = "flats.html"
    form_class = ApartmentForm
    permission_required = "role.has_apartments"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if "city" in self.request.GET:
            context["form"] = self.form_class(self.request.GET)
        else:
            context["form"] = self.form_class()

        return context


class FlatAjaxTable(AjaxDatatableView):
    model = Apartment
    title = "Квартиры"
    initial_order = [["number", "asc"]]

    def get_column_defs(self, request):
        house_choices = list(House.objects.all().values_list("id", "title"))
        return [
            {
                "name": "number",
                "title": "№ квартиры",
            },
            {
                "name": "house",
                "foreign_field": "house__title",
                "title": "Дом",
                "choices": house_choices,
            },
            {
                "name": "section",
                "foreign_field": "section__title",
                "title": "Секция",
                "visible": True,
                "searchable": False,
                "orderable": True,
            },
            {
                "name": "floor",
                "foreign_field": "floor__title",
                "title": "Этаж",
                "visible": True,
                "searchable": False,
                "orderable": True,
            },
            {
                "name": "owner",
                "title": "Владелец",
                "visible": True,
                "searchable": True,
                "orderable": True,
            },
            {
                "name": "remainder",
                "title": "Остаток (грн)",
                "visible": True,
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "actions",
                "visible": True,
                "searchable": False,
                "orderable": False,
                "title": "",
            },
        ]

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset()

        excluded_roles = [
            "Директор",
        ]
        if request.user.role.title not in excluded_roles:
            qs = qs.filter(house__owner=request.user)
            return qs

        house = self.request.POST.get("house")
        section = self.request.POST.get("section")
        floor = self.request.POST.get("floor")
        owner = self.request.POST.get("owner")

        if house:
            qs = qs.filter(house=house)
        if section:
            qs = qs.filter(section=section)
        if floor:
            qs = qs.filter(floor=floor)
        if owner:
            qs = qs.filter(owmer=owner)

        return qs

    def customize_row(self, row, obj):
        if obj.owner:
            row["owner"] = obj.owner.fullname

        edit_url = reverse("admin:edit-flat", args=[obj.id])
        delete_url = reverse("admin:delete-flat", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            edit_url,
            delete_url,
        )

        detail_url = reverse("admin:flat-detail", args=[obj.id])
        if not hasattr(obj, "bankbook") or obj.bankbook is None:
            row["remainder"] = "<span class='text-muted'>(нет счета)</span>"

        elif obj.bankbook:
            stats = CashBox.objects.filter(bank_book=obj.bankbook).aggregate(
                total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
                total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
            )

            receipts = Receipt.objects.filter(bankbook=obj.bankbook)

            receipts_pay = (
                ServiceFullCost.objects.filter(receipt__in=receipts).aggregate(
                    total=Sum("full_cost")
                )["total"]
                or 0
            )

            balance = (stats["total_cm"] or 0) - receipts_pay - (stats["total_og"] or 0)
            formatted_balance = "{:,.2f}".format(balance).replace(",", " ")
            if balance > 0:
                row["remainder"] = (
                    f'<span class="text-success">{formatted_balance}</span>'
                )
            elif balance < 0:
                row["remainder"] = (
                    f'<span class="text-danger">{formatted_balance}</span>'
                )
            else:
                row["remainder"] = "<span>0</span>"

        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}
        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

        return row


class EditFlat(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = Apartment
    template_name = "flat.html"
    form_class = ApartmentForm
    success_url = reverse_lazy("admin:flat-list")
    permission_required = "role.has_apartments"


class FlatDetail(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = Apartment
    template_name = "flat_detail.html"
    context_object_name = "apartment"
    permission_required = "role.has_apartments"


class DeleteFlat(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = Apartment
    permission_required = "role.has_apartments"

    def get(self, request, pk):
        flatt = get_object_or_404(Apartment, pk=pk)
        flatt.delete()
        return redirect("admin:flat-list")


class ListOwner(RedirectMixin, PermissionRequiredMixin, ListView):
    template_name = "owners.html"
    form_class = OwnerForm
    model = User
    context_object_name = "owner"
    permission_required = "role.has_owners"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()

        return context

    def get_queryset(self):
        return User.objects.filter(is_staff=False)


class DetailOwner(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = User
    template_name = "owner_detail.html"
    context_object_name = "owner"
    permission_required = "role.has_owners"


class SendInviteMessage(RedirectMixin, PermissionRequiredMixin, FormView):
    template_name = "owner_send_message.html"
    form_class = InviteForm
    permission_required = "role.has_owners"

    def form_valid(self, form):
        email = form.cleaned_data.get("email")
        send_invite_owner.delay(email)

        return HttpResponseRedirect(reverse_lazy("admin:owner-list"))


class CreateOwner(RedirectMixin, PermissionRequiredMixin, CreateView):
    template_name = "owner.html"
    form_class = OwnerForm
    model = User
    context_object_name = "owner"
    success_url = reverse_lazy("admin:flat-list")
    permission_required = "role.has_owners"

    def form_valid(self, form):
        user = form.save(commit=False)
        user.is_staff = False
        raw_password = form.cleaned_data.get("password1")
        user_email = form.cleaned_data.get("email")
        id = form.cleaned_data.get("id_user")
        if not id:
            try:
                last_box = User.objects.only("id").order_by("id").last()
                next_id = (last_box.id + 1) if last_box else 1
                user.id_user = "00" + str(next_id)
            except IntegrityError:
                pass
        user.save()
        send_email_verify(self.request, user)
        if len(raw_password) != 0:
            send_password.delay(raw_password, user_email)

        return HttpResponseRedirect(reverse_lazy("admin:owner-list"))


class EditOwner(RedirectMixin, PermissionRequiredMixin, UpdateView):
    template_name = "owner.html"
    form_class = OwnerForm
    model = User
    context_object_name = "owner"
    success_url = reverse_lazy("admin:owner-list")
    permission_required = "role.has_owners"

    def form_valid(self, form):
        user = form.save(commit=False)
        raw_password = form.cleaned_data.get("password1")
        user_email = form.cleaned_data.get("email")
        id = form.cleaned_data.get("id_user")

        if not id:
            last_box = User.objects.only("id").order_by("id").last()
            next_id = (last_box.id + 1) if last_box else 1
            user.id_user = "00" + str(next_id)

        user.save()
        if len(raw_password) != 0:
            send_password.delay(raw_password, user_email)

        return super().form_valid(form)


class SendVerificationView(RedirectMixin, PermissionRequiredMixin, View):
    model = User
    permission_required = "role.has_owners"

    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        send_email_verify(self.request, user)
        return redirect("admin:edit-owner", user.pk)


class OwnerAjaxTable(AjaxDatatableView):
    model = User
    title = "Квартиры"
    initial_order = [["full_name", "asc"]]

    def get_column_defs(self, request):
        return [
            {"name": "id_user", "title": "ID", "searchable": True, "orderable": False},
            {
                "name": "full_name",
                "title": "ФИО",
                "searchable": True,
                "orderable": True,
            },
            {
                "name": "phone_number",
                "title": "Телефон",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "email",
                "title": "E-mail",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "virtual_house",
                "title": "Дом",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "virtual_apartment",
                "title": "Квартира",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "date_joined",
                "title": "Добавлен",
                "searchable": True,
                "orderable": True,
            },
            {
                "name": "status",
                "title": "Статус",
                "searchable": True,
                "orderable": True,
                "choices": Status.choices,
            },
            {
                "name": "actions",
                "visible": True,
                "searchable": False,
                "orderable": False,
                "title": "",
            },
        ]

    def customize_row(self, row, obj):
        parts = [obj.second_name, obj.first_name, obj.last_name]
        row["full_name"] = " ".join(filter(None, parts))
        row["phone_number"] = str(obj.phone_number) if obj.phone_number else ""

        detail_url = reverse("admin:detail-owner", args=[obj.id])
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        edit_url = reverse("admin:edit-owner", args=[obj.id])
        delete_url = reverse("admin:delete-owner", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            edit_url,
            delete_url,
        )

        apartments = obj.apartment_set.all()

        if apartments:
            houses = set(a.house.title for a in apartments if a.house)
            row["virtual_house"] = ", ".join(houses)

            numbers = [str(a.number) for a in apartments if a.number]
            row["virtual_apartment"] = ", ".join(numbers)
        else:
            row["virtual_house"] = "(не задано)"
            row["virtual_apartment"] = "(не задано)"

        status_text = obj.get_status_display()

        if obj.status == Status.active:
            css_class = "label label-success"
        elif obj.status == Status.new:
            css_class = "label label-warning"
        elif obj.status == Status.inactive:
            css_class = "label label-danger"

        row["status"] = f'<small class="{css_class}">{status_text}</small>'

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

        return row

    def get_initial_queryset(self, request=None):
        qs = User.objects.filter(is_staff=False)

        qs = qs.annotate(
            full_name=Concat(
                "second_name",
                Value(" "),
                "first_name",
                Value(" "),
                "last_name",
                output_field=CharField(),
            )
        )

        qs = qs.annotate(
            virtual_house=Value("", output_field=CharField()),
            virtual_apartment=Value("", output_field=CharField()),
        )

        house_id = self.request.POST.get("house")
        apartment_id = self.request.POST.get("apartment")
        date_joined = self.request.POST.get("date_joined")

        if house_id:
            qs = qs.filter(apartment__house__id=house_id)

        if apartment_id:
            qs = qs.filter(apartment__id=apartment_id)

        if date_joined:
            qs = qs.filter(date_joined__date=date_joined)

        if house_id or apartment_id:
            qs = qs.distinct()

        return qs


class DeleteOwner(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = User
    permission_required = "role.has_owners"

    def get(self, request, pk):
        house = get_object_or_404(User, pk=pk)
        house.delete()
        return redirect("admin:owner-list")


class CreateBankBook(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = BankBook
    template_name = "bankbook.html"
    form_class = BankBookForm
    success_url = reverse_lazy("admin:bankbook-list")
    permission_required = "role.has_personal_accounts"


class BankBookListView(
    RedirectMixin, PermissionRequiredMixin, CommonContextMixin, ListView
):
    model = BankBook
    template_name = "bankbooks.html"
    context_object_name = "bankbook"
    form_class = BankBookForm
    permission_required = "role.has_personal_accounts"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()

        return context


class UpdateBankBook(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = BankBook
    template_name = "bankbook.html"
    form_class = BankBookForm
    context_object_name = "bankbook"
    success_url = reverse_lazy("admin:bankbook-list")
    permission_required = "role.has_personal_accounts"


def download_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "XLSX"

    columns = ["Лицевой счет", "Статус", "Дом", "Секция", "Квартира", "Владелец"]
    ws.append(columns)
    qs = BankBook.objects.select_related("house", "apartment", "section").all()

    for book in qs:
        row = [
            book.number,
            book.get_status_display()
            if hasattr(book, "get_status_display")
            else book.status,
            book.house.title if book.house else "",
            book.section.title if book.section else "",
            book.apartment.number if book.apartment else "",
            book.apartment.owner.fullname if book.apartment.owner else "",
        ]
        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bank_books.xlsx"'
    wb.save(response)

    return response


def export_status(request, task_id):
    result = AsyncResult(task_id)

    if result.ready():
        return JsonResponse({"ready": True, "file_url": result.result})

    return JsonResponse({"ready": False})


class DeleteBankBook(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = BankBook
    permission_required = "role.has_personal_accounts"

    def get(self, request, pk):
        bankbook = get_object_or_404(BankBook, pk=pk)
        bankbook.delete()
        return redirect("admin:bankbook-list")


class BankBookDetailView(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = BankBook
    template_name = "bankbook_detail.html"
    context_object_name = "bankbook"
    permission_required = "role.has_personal_accounts"

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        bankbook = self.object
        stats = CashBox.objects.filter(bank_book=bankbook).aggregate(
            total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
            total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
        )

        receipts = Receipt.objects.filter(bankbook=bankbook)

        receipts_pay = (
            ServiceFullCost.objects.filter(receipt__in=receipts).aggregate(
                total=Sum("full_cost")
            )["total"]
            or 0
        )
        balance = (stats["total_cm"] or 0) - receipts_pay - (stats["total_og"] or 0)
        formatted_balance = "{:,.2f}".format(balance).replace(",", " ")
        context["balance"] = formatted_balance

        return context


class BankBookAjaxTable(AjaxDatatableView):
    model = BankBook
    template_name = "bankbook.html"

    def get_column_defs(self, request=None):
        columns = [
            {"name": "number", "title": "№", "searchable": True},
            {
                "name": "status",
                "title": "Статус",
                "searchable": True,
                "choices": StatusBankBook.choices,
            },
            {
                "name": "apartment",
                "foreign_field": "apartment__number",
                "title": "Квартира",
                "searchable": True,
            },
            {
                "name": "house",
                "foreign_field": "house__title",
                "title": "Дом",
                "searchable": False,
            },
            {
                "name": "section",
                "foreign_field": "section__title",
                "title": "Секция",
                "searchable": False,
            },
            {
                "name": "owner",
                "foreign_field": "apartment__owner__last_name",
                "title": "Владелец",
                "searchable": False,
            },
            {
                "name": "remainder",
                "title": "Остаток (грн)",
                "searchable": False,
            },
            {
                "name": "actions",
                "title": "",
                "searchable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        status = obj.get_status_display()
        css_class = "label label-default"
        if obj.status == StatusBankBook.ACTIVE:
            css_class = "label label-success"
        elif obj.status == StatusBankBook.INACTIVE:
            css_class = "label label-danger"
        row["status"] = f'<small class="{css_class}">{status}</small>'

        detail_url = "#"
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        if obj.apartment and obj.apartment.owner:
            row["owner"] = obj.apartment.owner.fullname

        stats = CashBox.objects.filter(bank_book=obj).aggregate(
            total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
            total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
        )

        receipts = Receipt.objects.filter(bankbook=obj)

        receipts_pay = (
            ServiceFullCost.objects.filter(receipt__in=receipts).aggregate(
                total=Sum("full_cost")
            )["total"]
            or 0
        )

        balance = (stats["total_cm"] or 0) - receipts_pay - (stats["total_og"] or 0)
        formatted_balance = "{:,.2f}".format(balance).replace(",", " ")
        if balance > 0:
            row["remainder"] = f'<span class="text-success">{formatted_balance}</span>'
        elif balance < 0:
            row["remainder"] = f'<span class="text-danger">{formatted_balance}</span>'
        else:
            row["remainder"] = "<span>0</span>"

        edit_url = reverse("admin:bankbook-edit", args=[obj.id])
        delete_url = reverse("admin:bankbook-delete", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            edit_url,
            delete_url,
        )

        detail_url = reverse("admin:bankbook-detail", args=[obj.id])
        row["DT_RowAttr"] = row["DT_RowAttr"] = {
            "data-href": detail_url,
            "style": "cursor: pointer;",
        }

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

        return row

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset().order_by("id")

        house = self.request.POST.get("house")
        section = self.request.POST.get("section")
        owner = self.request.POST.get("owner")
        remainder = self.request.POST.get("remainder")

        if house:
            qs = qs.filter(house__id=house)
        if section:
            qs = qs.filter(section__id=section)
        if owner:
            qs = qs.filter(apartment__owner__fullname=owner)
        if remainder:
            if remainder == "remainder":
                qs = (
                    BankBook.objects.annotate(
                        total_cm=Coalesce(
                            Sum(
                                "cashbox__sum",
                                filter=Q(cashbox__status="CM", cashbox__is_catch=True),
                            ),
                            0.0,
                            output_field=FloatField(),
                        ),
                        total_og=Coalesce(
                            Sum(
                                "cashbox__sum",
                                filter=Q(cashbox__status="OG", cashbox__is_catch=True),
                            ),
                            0.0,
                            output_field=FloatField(),
                        ),
                        total_pay=Coalesce(
                            Sum("receipt__servicefullcost__full_cost"),
                            0.0,
                            output_field=FloatField(),
                        ),
                    )
                    .annotate(
                        final_balance=F("total_cm") - F("total_pay") - F("total_og")
                    )
                    .filter(final_balance__lt=0)
                )
            elif remainder == "not_remainder":
                qs = (
                    BankBook.objects.annotate(
                        total_cm=Coalesce(
                            Sum(
                                "cashbox__sum",
                                filter=Q(cashbox__status="CM", cashbox__is_catch=True),
                            ),
                            0.0,
                            output_field=FloatField(),
                        ),
                        total_og=Coalesce(
                            Sum(
                                "cashbox__sum",
                                filter=Q(cashbox__status="OG", cashbox__is_catch=True),
                            ),
                            0.0,
                            output_field=FloatField(),
                        ),
                        total_pay=Coalesce(
                            Sum("receipt__servicefullcost__full_cost"),
                            0.0,
                            output_field=FloatField(),
                        ),
                    )
                    .annotate(
                        final_balance=F("total_cm") - F("total_pay") - F("total_og")
                    )
                    .filter(final_balance__gte=0)
                )

        return qs


class CreateCounter(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = Counter
    form_class = CounterForm
    template_name = "counter.html"
    success_url = reverse_lazy("admin:counter-list")
    permission_required = "role.has_meters"

    def get_initial(self):
        initial = super().get_initial()
        source_id = self.request.GET.get("source_id")

        if source_id:
            try:
                original = Counter.objects.get(pk=source_id)
                initial.update(
                    {
                        "house": original.house,
                        "section": original.section,
                        "apartment": original.apartment,
                        "service": original.service,
                    }
                )
            except Counter.DoesNotExist:
                pass
        return initial


class CounterList(RedirectMixin, PermissionRequiredMixin, ListView):
    model = Counter
    form_class = CounterForm
    template_name = "counters.html"
    context_object_name = "counters"
    permission_required = "role.has_meters"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()

        return context


class CounterAjaxTable(AjaxDatatableView):
    model = Counter
    title = "Счетчики"
    initial_order = [[2, "asc"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "house",
                "title": "Дом",
                "foreign_field": "house__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "section",
                "title": "Секция",
                "foreign_field": "section__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "apartment",
                "title": "№ квартиры",
                "foreign_field": "apartment__number",
                "searchable": True,
                "orderable": True,
            },
            {
                "name": "service",
                "title": "Счетчик",
                "foreign_field": "service__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "readings",
                "title": "Текущие показания",
                "searchable": False,
                "orderable": False,
                "className": "rows",
            },
            {
                "name": "units",
                "foreign_field": "service__units_of_measure__units",
                "title": "Текущие показания",
                "className": "rows",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "actions",
                "searchable": False,
                "orderable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        detail_url = (
            reverse("admin:counter-specific-list") + f"?apartment_id={obj.apartment.id}"
        )
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        duplicate_url = reverse("admin:counter-create") + f"?source_id={obj.pk}"
        story_url = (
            reverse("admin:counter-specific-list") + f"?apartment_id={obj.apartment.id}"
        )

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-dashboard"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-eye"></i></a>'
            "</div>",
            duplicate_url,
            story_url,
        )

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

        return row

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset().order_by("id")

        house = self.request.POST.get("house")
        section = self.request.POST.get("section")
        service = self.request.POST.get("service")
        if house:
            qs = qs.filter(house__id=house)
        if section:
            qs = qs.filter(section__id=section)
        if service:
            qs = qs.filter(service__id=service)

        return qs


class CounterSpecificList(RedirectMixin, PermissionRequiredMixin, ListView):
    model = Counter
    form_class = CounterForm
    template_name = "counter_specific.html"
    permission_required = "role.has_meters"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()

        return context


class CounterSpecificAjaxTable(AjaxDatatableView):
    model = Counter
    title = "Счетчики"
    initial_order = [[2, "des"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "number",
                "title": "№",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "status",
                "title": "Статус",
                "searchable": True,
                "orderable": False,
                "choices": StatusCounter.choices,
            },
            {
                "name": "date",
                "title": "Дата",
                "orderable": True,
                "searchable": True,
            },
            {
                "name": "service",
                "foreign_field": "service__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "readings",
                "title": "Показания",
                "searchable": False,
                "orderable": False,
                "className": "rows",
            },
            {
                "name": "units",
                "foreign_field": "service__units_of_measure__units",
                "searchable": False,
                "orderable": False,
                "className": "rows",
            },
            {
                "name": "actions",
                "searchable": False,
                "orderable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        status = obj.get_status_display()
        css_class = "label label-default"
        if obj.status == StatusCounter.NEW:
            css_class = "label label-warning"
        elif obj.status == StatusCounter.TAKEN:
            css_class = "label label-success"
        elif obj.status == StatusCounter.TAKEN_AND_PAID:
            css_class = "label label-success"
        elif obj.status == StatusCounter.NULLABLE:
            css_class = "label label-primary"
        row["status"] = f'<small class="{css_class}">{status}</small>'

        detail_url = reverse("admin:counter-detail", args=[obj.id])
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        edit_url = reverse("admin:counter-edit", args=[obj.id])
        delete_url = reverse("admin:counter-delete", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            edit_url,
            delete_url,
        )

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'
        return row

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset(request)
        apartment_id = (
            self.request.POST.get("apartment_id")
            or self.request.GET.get("apartment_id")
            or self.request.POST.get("extra_data[apartment_id]")
        )
        return qs.filter(apartment_id=apartment_id)


class CounterDetail(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = Counter
    template_name = "counter_detail.html"
    context_object_name = "counter"
    permission_required = "role.has_meters"


class CounterEdit(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = Counter
    form_class = CounterForm
    template_name = "counter.html"
    success_url = reverse_lazy("admin:counter-list")
    context_object_name = "counter"
    permission_required = "role.has_meters"


class DeleteCounter(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = Counter
    permission_required = "role.has_meters"

    def get(self, request, pk):
        counter = get_object_or_404(Counter, pk=pk)
        counter.delete()
        return redirect("admin:counter-list")


class CreateMasterCall(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = MasterCall
    form_class = MasterCallForm
    template_name = "master_call.html"
    success_url = reverse_lazy("admin:statistic")
    permission_required = "role.has_master_requests"


class MasterCallDetail(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = MasterCall
    template_name = "master_call_detail.html"
    context_object_name = "master_call"
    permission_required = "role.has_master_requests"


class MasterCallList(RedirectMixin, PermissionRequiredMixin, ListView):
    model = MasterCall
    form_class = MasterCallForm
    template_name = "master_calls.html"
    permission_required = "role.has_master_requests"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()
        masters_type = Role.objects.exclude(
            title__in=["Директор", "Управляющий", "Бухгалтер"]
        )
        context["master_types"] = masters_type
        return context


class MasterCallAjaxDataTable(AjaxDatatableView):
    model = MasterCall
    title = "Заявки вызова мастера"
    initial_order = [[0, "desc"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "id",
                "title": "№ заявки",
                "orderable": True,
                "searchable": True,
            },
            {
                "name": "date",
                "title": "Удобное время",
                "orderable": True,
                "searchable": True,
            },
            {
                "name": "master_type",
                "title": "Тип мастера",
                "orderable": True,
                "searchable": False,
            },
            {
                "name": "description",
                "title": "Описание",
                "orderable": False,
                "searchable": True,
            },
            {
                "name": "apartment",
                "title": "Квартира",
                "orderable": False,
                "searchable": True,
            },
            {
                "name": "owner",
                "title": "Владелец",
                "orderable": False,
                "searchable": True,
            },
            {
                "name": "phone_number",
                "foreign_field": "master__phone_number",
                "title": "Телефон",
                "orderable": False,
                "searchable": True,
            },
            {
                "name": "master",
                "title": "Мастер",
                "orderable": False,
                "searchable": False,
            },
            {
                "name": "status",
                "title": "Статус",
                "orderable": False,
                "searchable": True,
                "choices": StatusCall.choices,
            },
            {
                "name": "actions",
                "searchable": False,
                "orderable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        apartment_url = reverse("admin:flat-detail", args=[obj.apartment.id])
        row["apartment"] = format_html(
            '<a href="{}">Кв.{}, {} </a>',
            apartment_url,
            obj.apartment.number,
            obj.apartment.house.title,
        )
        row["master_type"] = (
            obj.master_type.title if obj.master_type else "Любой специалист"
        )
        if obj.owner:
            owner_url = reverse("admin:detail-owner", args=[obj.owner.id])
            row["owner"] = format_html(
                '<a href="{}">{}</a>', owner_url, obj.owner.fullname
            )

        if obj.master:
            master_url = reverse("settings:user-detail", args=[obj.master.id])
            row["master"] = format_html(
                '<a href="{}">{}-{}</a>',
                master_url,
                obj.master.role.title,
                obj.master.fullname,
            )

        css_class = "label label-default"
        if obj.call_status == StatusCall.IN_WORK:
            css_class = "label label-warning"
        elif obj.call_status == StatusCall.COMPLETED:
            css_class = "label label-success"
        elif obj.call_status == StatusCall.NEW:
            css_class = "label label-primary"
        row["status"] = f'<small class="{css_class}">{obj.call_status}</small>'

        detail_url = reverse("admin:master-call-detail", args=[obj.id])
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        edit_url = reverse("admin:master-call-edit", args=[obj.id])
        delete_url = reverse("admin:master-call-delete", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            edit_url,
            delete_url,
        )

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'
        return row

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset().order_by("id")
        excluded_roles = ["Директор", "Управляющий", "Бухгалтер"]
        if request.user.role.title not in excluded_roles:
            qs = qs.filter(master=request.user)

        master = self.request.POST.get("master")
        date_from = self.request.POST.get("date_from")
        date_to = self.request.POST.get("date_to")
        master_type = self.request.POST.get("master_type")

        if master:
            qs = qs.filter(master_id=master)

        if date_to:
            qs = qs.filter(date__lte=date_to)

        if date_from:
            qs = qs.filter(date__gte=date_from)

        if master_type:
            if master_type == "noone":
                qs = qs.filter(master_type__isnull=True)
            else:
                qs = qs.filter(master_type=master_type)

        return qs


class EditMasterCall(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = MasterCall
    form_class = MasterCallForm
    template_name = "master_call.html"
    success_url = reverse_lazy("admin:master-call-list")
    context_object_name = "mastercall"
    permission_required = "role.has_master_requests"


class DeleteMasterCall(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = MasterCall
    permission_required = "role.has_master_requests"

    def get(self, request, pk):
        master_call = get_object_or_404(MasterCall, pk=pk)
        master_call.delete()
        return redirect("admin:master-call-list")


class ReceiptList(RedirectMixin, PermissionRequiredMixin, CommonContextMixin, ListView):
    model = Receipt
    template_name = "receipts.html"
    form_class = ReceiptForm
    permission_required = "role.has_invoices"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = self.form_class()
        return context


class ReceiptCreate(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = Receipt
    form_class = ReceiptForm
    template_name = "receipt.html"
    success_url = reverse_lazy("admin:receipt-list")
    permission_required = "role.has_invoices"

    def get_duplicate(self):
        duplicate_id = self.request.GET.get("duplicate")
        if not duplicate_id:
            return None
        try:
            return Receipt.objects.get(pk=duplicate_id)
        except Receipt.DoesNotExist:
            return None

    def get_source_bankbook(self):
        source_id = self.request.GET.get("source_id")
        if not source_id:
            return None
        try:
            return BankBook.objects.get(pk=source_id)
        except BankBook.DoesNotExist:
            return None

    def get_initial(self):
        initial = super().get_initial()

        duplicate = self.get_duplicate()
        bankbook = self.get_source_bankbook()

        if duplicate:
            initial.update(
                {
                    "house": duplicate.house_id,
                    "section": duplicate.section_id,
                    "apartment": duplicate.apartment_id,
                    "tariff": duplicate.tariff_id,
                    "bankbook": duplicate.bankbook.number
                    if duplicate.bankbook
                    else None,
                }
            )

        if bankbook:
            initial.update(
                {
                    "house": bankbook.house_id,
                    "section": bankbook.section_id,
                    "apartment": bankbook.apartment_id,
                    "bankbook": bankbook,
                }
            )

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        duplicate = self.get_duplicate()

        if self.request.POST:
            context["services"] = ServiceFullCostFormSet(
                self.request.POST, prefix="services"
            )
        else:
            if duplicate:
                initial_data = [
                    {
                        "service": s.service_id,
                        "cost": s.cost,
                        "unit": s.unit_id,
                        "full_cost": s.full_cost,
                        "consumption": s.consumption,
                    }
                    for s in duplicate.servicefullcost_set.all()
                ]

                DuplicateFormSet = inlineformset_factory(
                    Receipt,
                    ServiceFullCost,
                    form=ServiceFullCostForm,
                    extra=len(initial_data),
                    can_delete=True,
                )

                context["services"] = DuplicateFormSet(
                    instance=None,
                    prefix="services",
                    initial=initial_data,
                    queryset=ServiceFullCost.objects.none(),
                )
            else:
                context["services"] = ServiceFullCostFormSet(
                    instance=None, prefix="services"
                )

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        services = context["services"]

        if not services.is_valid():
            return self.form_invalid(form)

        duplicate = self.get_duplicate()

        with transaction.atomic():
            self.object = form.save(commit=False)

            if duplicate:
                self.object.source_id = duplicate.id

            counter_ids_str = self.request.POST.get("counter_ids")
            if counter_ids_str:
                try:
                    self.object.counter_ids = json.loads(counter_ids_str)
                except (ValueError, TypeError):
                    self.object.counter_ids = []

            bank_book = form.cleaned_data.get("bankbook")
            if bank_book:
                stats = CashBox.objects.filter(bank_book=bank_book).aggregate(
                    total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
                    total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
                )

                receipts_pay = (
                    ServiceFullCost.objects.filter(
                        receipt__bankbook=bank_book
                    ).aggregate(total=Sum("full_cost"))["total"]
                    or 0
                )

                balance = (
                    (stats["total_cm"] or 0) - (stats["total_og"] or 0) - receipts_pay
                )

                requested_sum = form.cleaned_data.get("sum") or 0

                if balance >= requested_sum:
                    self.object.status = "PD"

            self.object.save()
            services.instance = self.object
            services.save()

        return HttpResponseRedirect(self.get_success_url())

    def form_invalid(self, form):
        print(form.errors)
        return self.render_to_response(self.get_context_data(form=form))


def get_tariff(request):
    apartment_id = request.GET.get("apartment_id")
    data = {"tariff_id": None}
    if apartment_id:
        try:
            tariff = Tariffs.objects.filter(apartment=apartment_id).first()
            if tariff:
                data["tariff_id"] = tariff.id
        except Exception as e:
            print(f"Error in get_tariff: {e}")
    return JsonResponse(data)


def get_tariff_services(request):
    tariff_id = request.GET.get("tariff_id")

    if not tariff_id:
        return JsonResponse({"success": False, "error": "Не указан тариф"})

    try:
        tariff = Tariffs.objects.get(id=tariff_id)
        services_cost = ServicesCost.objects.filter(tariff=tariff).select_related(
            "service", "service__units_of_measure"
        )

        services_data = []
        for sc in services_cost:
            service_info = {
                "service_id": sc.service.id,
                "service_name": sc.service.name
                if hasattr(sc.service, "name")
                else str(sc.service),
                "unit_id": sc.service.units_of_measure.id
                if sc.service.units_of_measure
                else None,
                "unit_name": sc.service.units_of_measure.units
                if sc.service.units_of_measure
                else "",
                "cost": str(sc.cost),
            }
            services_data.append(service_info)

        return JsonResponse({"success": True, "services": services_data})

    except Tariffs.DoesNotExist:
        return JsonResponse({"success": False, "error": "Тариф не найден"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def get_unit(request):
    service_id = request.GET.get("service_id")
    data = {"unit_id": None}
    if service_id:
        try:
            service = Service.objects.get(id=service_id)
            if service.units_of_measure:
                data["unit_id"] = service.units_of_measure.id
        except (Service.DoesNotExist, AttributeError):
            pass
    return JsonResponse(data)


def get_unit_service(request):
    service_id = request.GET.get("service_id")
    data = {"unit_id": None, "cost": None}
    if service_id:
        try:
            service = Service.objects.get(id=service_id)
            if service.units_of_measure:
                data["unit_id"] = service.units_of_measure.id
            if hasattr(service, "price") and service.price:
                data["cost"] = str(service.price)
        except (Service.DoesNotExist, AttributeError):
            pass
    return JsonResponse(data)


def get_counter_readings(request):
    apartment_id = request.GET.get("apartment_id")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if not apartment_id:
        return JsonResponse({"success": False, "error": "Не указана квартира"})

    try:
        counters_query = Counter.objects.filter(apartment_id=apartment_id, status="NEW")

        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                counters_query = counters_query.filter(date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                counters_query = counters_query.filter(date__lte=date_to_obj)
            except ValueError:
                pass

        if not counters_query.exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": "Для выбранной квартиры не найдено новых счетчиков в указанном периоде",
                }
            )

        counter_readings = []
        services = counters_query.values_list("service", flat=True).distinct()

        for service_id in services:
            service_counters = counters_query.filter(service_id=service_id)
            total_consumption = (
                service_counters.aggregate(total=Sum("readings"))["total"] or 0
            )
            first_counter = service_counters.first()

            if first_counter:
                service = first_counter.service
                counter_ids = list(service_counters.values_list("id", flat=True))

                reading_info = {
                    "counter_ids": counter_ids,
                    "service_id": service.id,
                    "service_name": service.name
                    if hasattr(service, "name")
                    else str(service),
                    "unit_id": service.units_of_measure.id
                    if hasattr(service, "units_of_measure") and service.units_of_measure
                    else None,
                    "consumption": str(total_consumption),
                    "count": service_counters.count(),
                }
                counter_readings.append(reading_info)

        return JsonResponse({"success": True, "readings": counter_readings})

    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Ошибка при получении показаний: {str(e)}"}
        )


def mark_counters_taken(request):
    try:
        data = json.loads(request.body)
        counter_ids = data.get("counter_ids", [])

        if not counter_ids:
            return JsonResponse({"success": False, "error": "Не указаны ID счетчиков"})

        updated_count = Counter.objects.filter(id__in=counter_ids, status="NEW").update(
            status="TAKEN"
        )

        return JsonResponse({"success": True, "updated_count": updated_count})

    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Ошибка при обновлении статусов: {str(e)}"}
        )


def mark_counters_paid(request):
    try:
        data = json.loads(request.body)
        invoice_id = data.get("invoice_id")

        if not invoice_id:
            return JsonResponse({"success": False, "error": "Не указан ID квитанции"})

        try:
            invoice = Receipt.objects.get(id=invoice_id)
        except Receipt.DoesNotExist:
            return JsonResponse({"success": False, "error": "Квитанция не найдена"})

        if invoice.status != "PD":
            return JsonResponse(
                {"success": False, "error": 'Квитанция должна иметь статус "Оплачена"'}
            )

        counter_ids = invoice.counter_ids if invoice.counter_ids else []

        if not counter_ids:
            return JsonResponse(
                {"success": False, "error": "У квитанции нет связанных счетчиков"}
            )

        updated_count = Counter.objects.filter(
            id__in=counter_ids, status="TAKEN"
        ).update(status="TAKEN_AND_PAID")

        return JsonResponse(
            {"success": True, "updated_count": updated_count, "invoice_id": invoice_id}
        )

    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Ошибка при обновлении статусов: {str(e)}"}
        )


class ReceiptCounterTable(AjaxDatatableView):
    model = Counter
    initial_order = [[0, "asc"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "number",
                "title": "№",
                "searchable": False,
                "orderable": True,
            },
            {
                "name": "status",
                "title": "Статус",
                "searchable": False,
                "orderable": True,
            },
            {
                "name": "date",
                "title": "Дата",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "house",
                "title": "Дом",
                "foreign_field": "house__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "section",
                "title": "Секция",
                "foreign_field": "section__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "apartment",
                "title": "№ квартиры",
                "foreign_field": "apartment__number",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "service",
                "title": "Счетчик",
                "foreign_field": "service__title",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "readings",
                "title": "Текущие показания",
                "searchable": False,
                "orderable": False,
                "className": "rows",
            },
            {
                "name": "units",
                "foreign_field": "service__units_of_measure__units",
                "title": "Текущие показания",
                "className": "rows",
                "searchable": False,
                "orderable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        css_class = "label label-default"
        if obj.status == StatusCounter.NEW:
            css_class = "label label-warning"
        elif obj.status == StatusCounter.TAKEN:
            css_class = "label label-success"
        elif obj.status == StatusCounter.TAKEN_AND_PAID:
            css_class = "label label-success"
        row["status"] = f'<small class="{css_class}">{obj.get_status_display()}</small>'

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

        return row

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset().order_by("id")

        house = self.request.POST.get("house")
        apartment = self.request.POST.get("apartment")
        date_from = self.request.POST.get("date_from")
        date_to = self.request.POST.get("date_to")

        if house:
            qs = qs.filter(house__id=house)

        if apartment:
            qs = qs.filter(apartment__id=apartment)

        if date_to:
            qs = qs.filter(date__lte=date_to)

        if date_from:
            qs = qs.filter(date__gte=date_from)

        return qs


class EditReceipt(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = Receipt
    form_class = ReceiptForm
    template_name = "receipt.html"
    success_url = reverse_lazy("admin:receipt-list")
    context_object_name = "receipt"
    permission_required = "role.has_invoices"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["services"] = ServiceFullCostFormSet(
                self.request.POST, instance=self.object, prefix="services"
            )
        else:
            context["services"] = ServiceFullCostFormSet(
                instance=self.object, prefix="services"
            )
        return context

    def form_valid(self, form):
        services = ServiceFullCostFormSet(
            self.request.POST, instance=self.object, prefix="services"
        )

        if not services.is_valid():
            return self.form_invalid(form)

        with transaction.atomic():
            self.object = form.save(commit=False)

            counter_ids_str = self.request.POST.get("counter_ids")
            if counter_ids_str:
                try:
                    self.object.counter_ids = json.loads(counter_ids_str)
                except (ValueError, TypeError):
                    self.object.counter_ids = []

            bank_book = form.cleaned_data.get("bankbook")
            if bank_book:
                stats = CashBox.objects.filter(bank_book=bank_book).aggregate(
                    total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
                    total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
                )

                receipts_pay = (
                    ServiceFullCost.objects.filter(
                        receipt__bankbook=bank_book
                    ).aggregate(total=Sum("full_cost"))["total"]
                    or 0
                )

                balance = (
                    (stats["total_cm"] or 0) - (stats["total_og"] or 0) - receipts_pay
                )

                requested_sum = form.cleaned_data.get("sum") or 0

                if balance >= requested_sum:
                    self.object.status = "PD"

            self.object.save()
            services.save()

        return HttpResponseRedirect(self.get_success_url())


class DeleteReceipt(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = Receipt
    permission_required = "role.has_invoices"

    def get(self, request, pk):
        receipt = get_object_or_404(Receipt, pk=pk)
        receipt.delete()
        return redirect("admin:receipt-list")


def receipt_bulk_delete(request):
    if request.method == "POST":
        item_ids = request.POST.getlist("item_ids")
        if item_ids:
            Receipt.objects.filter(id__in=item_ids).delete()

    return redirect("admin:receipt-list")


class ReceiptAjaxDatatable(AjaxDatatableView):
    model = Receipt
    initial_order = [[3, "asc"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "checkbox",
                "title": "<input type='checkbox' class='form-control-input checkbox-toggle'>",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "number",
                "title": "№ квитанции",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "status",
                "title": "Статус",
                "searchable": True,
                "orderable": False,
                "choices": StatusReceipt.choices,
            },
            {
                "name": "date",
                "title": "Дата",
                "searchable": False,
                "orderable": True,
            },
            {
                "name": "apartment",
                "title": "Квартира",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "owner",
                "title": "Владелец",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "is_catch",
                "title": "Проведена",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "sum",
                "title": "Сумма (грн)",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "actions",
                "title": "",
                "searchable": False,
                "orderable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        row["checkbox"] = mark_safe(
            f'<input type="checkbox" name="item_ids" value="{obj.id}" class="item-checkbox">'
        )
        row["apartment"] = (
            f"<small>{obj.apartment.number}, {obj.apartment.house.title}</small>"
        )
        row["owner"] = f"<small>{obj.apartment.owner.fullname}</small>"

        css_class = ""
        if obj.status == StatusReceipt.PAID:
            css_class = "label label-success"
        elif obj.status == StatusReceipt.PART:
            css_class = "label label-warning"
        elif obj.status == StatusReceipt.UNPAID:
            css_class = "label label-danger"
        row["status"] = f'<small class="{css_class}">{obj.get_status_display()}</small>'

        services_cost = obj.servicefullcost_set.all()
        services_sum = sum(a.full_cost for a in services_cost)
        row["sum"] = f"<small>{services_sum}</small>"

        if obj.is_catch:
            catch = "<small>Проведено</small>"
        elif not obj.is_catch:
            catch = "<small>Не проведено</small>"

        row["is_catch"] = catch

        duplicate_url = reverse("admin:receipt-create") + f"?duplicate={obj.pk}"
        edit_url = reverse("admin:receipt-edit", args=[obj.id])
        delete_url = reverse("admin:receipt-delete", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-clone"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            duplicate_url,
            edit_url,
            delete_url,
        )

        detail_url = reverse("admin:receipt-detail", args=[obj.id])
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset()
        bank_book_id = (
            self.request.POST.get("bank_book_id")
            or self.request.GET.get("bank_book_id")
            or self.request.POST.get("extra_data[bank_book_id]")
        )

        apartment_id = (
            self.request.POST.get("apartment_id")
            or self.request.GET.get("apartment_id")
            or self.request.POST.get("extra_data[apartment_id]")
        )

        owner = self.request.POST.get("owner")

        if apartment_id:
            qs = qs.filter(apartment_id=apartment_id)

        if bank_book_id:
            qs = qs.filter(bankbook_id=bank_book_id)

        if owner:
            qs = qs.filter(apartment__owner=owner)

        return qs


class DetailReceipt(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = Receipt
    template_name = "receipt_detail.html"
    context_object_name = "receipt"
    permission_required = "role.has_invoices"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receipt = self.object
        services_cost = receipt.servicefullcost_set.all()
        context["fullcost"] = sum(a.full_cost for a in services_cost)

        return context


class XlsTemplateReceiptDetail(
    RedirectMixin, PermissionRequiredMixin, FormMixin, DetailView
):
    model = Receipt
    template_name = "receipt_templates.html"
    form_class = XlsTemplateShowForm
    permission_required = "role.has_invoices"

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context["form"] = self.form_class()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        template_id = request.POST.get("templates")

        if "action_download" in request.POST:
            return self.handle_download(request, template_id, self.object.pk)

        elif "action_send_email" in request.POST:
            return self.handle_send(request, template_id, self.object.pk)

        return super().post(request, *args, **kwargs)

    def handle_download(self, request, template_id, pk):
        return download_receipt_xls(self.request, template_id, pk)

    def handle_send(self, request, template_id, pk):
        return send_receipt(self.request, template_id, pk)


def download_receipt_xls(request, template_id, pk):
    task = generate_excel_task.delay(template_id, pk)
    return JsonResponse({"task_id": task.id})


def send_receipt(request, template_id, pk):
    task = send_pdf_task.delay(template_id, pk)
    return JsonResponse({"task_id": task.id})


class XlsTemplateSettings(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = XlsTemplate
    form_class = XlsTemplateForm
    template_name = "receipt_trmplates_setting.html"
    success_url = reverse_lazy("admin:receipt-templates-settings")
    permission_required = "role.has_invoices"

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context["templates"] = XlsTemplate.objects.all()
        return context

    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)


class XlsTemplateSetDefault(RedirectMixin, PermissionRequiredMixin, View):
    permission_required = "role.has_invoices"

    def get(self, request, pk):
        template = get_object_or_404(XlsTemplate, pk=pk)
        template.set_as_default()
        return redirect("admin:receipt-templates-settings")


class XlsTemplateDelete(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = XlsTemplate
    permission_required = "role.has_invoices"

    def get(self, request, pk):
        templa = get_object_or_404(XlsTemplate, pk=pk)
        templa.delete()
        return redirect("admin:receipt-list")


class MessageAjaxDatatable(AjaxDatatableView):
    model = Message
    initial_order = [[2, "asc"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "checkbox",
                "title": "",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "recipient",
                "title": "Получатели",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "theme",
                "title": "Текст",
                "searchable": True,
                "orderable": True,
            },
            {
                "name": "created_at",
                "title": "Дата",
                "searchable": False,
                "orderable": False,
            },
        ]
        return columns

    def customize_row(self, row, obj):
        row["checkbox"] = mark_safe(
            f'<input type="checkbox" name="item_ids" value="{obj.id}" class="item-checkbox">'
        )

        if not obj.house:
            row["recipient"] = '<span><a href="#">Всем</a></span>'
        else:
            parts = [obj.house.title]

            if obj.section:
                parts.append(obj.section.title)
            if obj.floor:
                parts.append(obj.floor.title)
            if obj.apartment:
                parts.append(f"кв.{obj.apartment.number}")

            address_str = ", ".join(parts)
            row["recipient"] = f'<span><a href="#">{address_str}</a></span>'

        row["theme"] = f"<span>{obj.theme}-{obj.message}</span>"

        detail_url = reverse("admin:message-detail", args=[obj.id])

        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'


class CreateMessage(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = Message
    form_class = MessageForm
    template_name = "message.html"
    success_url = reverse_lazy("admin:message-list")
    permission_required = "role.has_messages"

    def get_initial(self):
        initial = super().get_initial()
        if self.request.GET.get("for_debtor") == "true":
            initial["for_debtor"] = True
        return initial

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.sender = self.request.user
        self.object.save()
        users = User.objects.filter(is_staff=False)
        if form.cleaned_data["house"]:
            users = users.filter(house_set=form.cleaned_data["house"])
        if form.cleaned_data["section"]:
            users = users.filter(section_set=form.cleaned_data["section"])
        if form.cleaned_data["floor"]:
            users = users.filter(floor_set=form.cleaned_data["floor"])
        if form.cleaned_data["apartment"]:
            users = users.filter(apartment=form.cleaned_data["apartment"])
        if form.cleaned_data["for_debtor"]:
            target_status = ["UPD", "PT"]
            users = users.filter(apartment__receipt__status__in=target_status)

        user_messages = [
            MessageStatus(user=user, message=self.object) for user in users
        ]
        MessageStatus.objects.bulk_create(user_messages)

        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)


class MessageDelete(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = Message
    permission_required = "role.has_messages"

    def get(self, request, pk):
        message = get_object_or_404(Message, pk=pk)
        message.delete()
        return redirect("admin:message-list")


class MessageList(RedirectMixin, PermissionRequiredMixin, ListView):
    model = Message
    template_name = "messages.html"
    permission_required = "role.has_messages"


class MessageDetailView(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = Message
    template_name = "message_detail.html"
    context_object_name = "message"
    permission_required = "role.has_messages"


def message_bulk_delete(request):
    if request.method == "POST":
        item_ids = request.POST.getlist("item_ids")
        if item_ids:
            Message.objects.filter(id__in=item_ids).delete()

    return redirect("admin:message-list")


class CashBoxComingCreateView(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = CashBox
    template_name = "cashbox_coming_add.html"
    form_class = CashBoxForm
    success_url = reverse_lazy("admin:cashbox-list")
    permission_required = "role.has_cashbox"

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.status = "CM"
        self.object.save()
        return super().form_valid(form)

    def get_initial(self):
        initial = super().get_initial()
        source_id = self.request.GET.get("source_id")
        bankbook_id = self.request.GET.get("bankbook_id")
        if source_id:
            try:
                original = CashBox.objects.get(pk=source_id)
                initial.update(
                    {
                        "article": original.article,
                        "is_catch": original.is_catch,
                        "comment": original.comment,
                        "owner": original.owner,
                        "bank_book": original.bank_book,
                    }
                )
            except CashBox.DoesNotExist:
                pass
        elif bankbook_id:
            try:
                original = BankBook.objects.get(pk=bankbook_id)
                initial.update(
                    {
                        "owner": original.apartment.owner,
                        "bank_book": original,
                    }
                )
            except CashBox.DoesNotExist:
                pass

        initial["manager"] = self.request.user
        return initial


class CashBoxOutGoCreateView(RedirectMixin, PermissionRequiredMixin, CreateView):
    model = CashBox
    template_name = "cashbox_outgo_add.html"
    form_class = CashBoxForm
    success_url = reverse_lazy("admin:cashbox-list")
    permission_required = "role.has_cashbox"

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.status = "OG"

        self.object.save()
        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)

    def get_initial(self):
        initial = super().get_initial()
        source_id = self.request.GET.get("source_id")
        if source_id:
            try:
                original = CashBox.objects.get(pk=source_id)
                initial.update(
                    {
                        "article": original.article,
                        "is_catch": original.is_catch,
                        "manager": original.manager,
                        "comment": original.comment,
                    }
                )
            except CashBox.DoesNotExist:
                pass
        else:
            initial["manager"] = self.request.user
        return initial


class CashBoxComingEditView(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = CashBox
    template_name = "cashbox_coming_add.html"
    form_class = CashBoxForm
    success_url = reverse_lazy("admin:cashbox-list")
    context_object_name = "cashbox"
    permission_required = "role.has_cashbox"

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.status = "CM"

        self.object.save()
        return super().form_valid(form)


class CashBoxOutGoEditView(RedirectMixin, PermissionRequiredMixin, UpdateView):
    model = CashBox
    template_name = "cashbox_outgo_add.html"
    form_class = CashBoxForm
    success_url = reverse_lazy("admin:cashbox-list")
    context_object_name = "cashbox"
    permission_required = "role.has_cashbox"

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.status = "OG"

        self.object.save()
        return super().form_valid(form)


class CashBoxList(CommonContextMixin, RedirectMixin, PermissionRequiredMixin, ListView):
    model = CashBox
    template_name = "cashbox_list.html"
    form_class = CashBoxForm
    permission_required = "role.has_cashbox"

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        stats = CashBox.objects.aggregate(
            total_cm=Sum("sum", filter=Q(status="CM", is_catch=True)),
            total_og=Sum("sum", filter=Q(status="OG", is_catch=True)),
        )
        formatted_comings = "{:,.2f}".format(stats["total_cm"] or 0).replace(",", " ")
        formatted_outgoes = "{:,.2f}".format(stats["total_og"] or 0).replace(",", " ")
        context["comings"] = formatted_comings
        context["outgoes"] = formatted_outgoes
        context["form"] = self.form_class
        return context


class CashBoxAjaxDatatable(AjaxDatatableView):
    model = CashBox
    initial_order = [[1, "desc"]]

    def get_column_defs(self, request):
        columns = [
            {
                "name": "number",
                "title": "№",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "date",
                "title": "Дата",
                "searchable": True,
                "orderable": True,
            },
            {
                "name": "statuss",
                "title": "Статус",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "article",
                "title": "Тип платежа",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "owner",
                "title": "Владелец",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "bankbook",
                "title": "Лицевой счет",
                "foreign_field": "bank_book__number",
                "searchable": True,
                "orderable": False,
            },
            {
                "name": "status",
                "title": "Приход/расход",
                "searchable": True,
                "orderable": False,
                "choices": StatusCashBox.choices,
            },
            {
                "name": "sum",
                "title": "Сумма (грн)",
                "searchable": False,
                "orderable": False,
            },
            {
                "name": "actions",
                "title": "",
                "searchable": False,
                "orderable": False,
            },
        ]

        return columns

    def customize_row(self, row, obj):
        row["date"] = obj.date.strftime("%d.%m.%Y")
        row["statuss"] = "Проведен" if obj.is_catch else "Не проведен"
        css_class = ""
        if obj.status == StatusCashBox.COMING:
            css_class = "text-success"
            row["sum"] = f'<small class="text-success">{obj.sum}</small>'
        elif obj.status == StatusCashBox.OUTGO:
            css_class = "text-danger"
            row["sum"] = f'<small class="text-danger">-{obj.sum}</small>'
        row["status"] = f'<small class="{css_class}">{obj.get_status_display()}</small>'
        detail_url = reverse("admin:cashbox-detail", args=[obj.id])
        row["DT_RowAttr"] = {"data-href": detail_url, "style": "cursor: pointer;"}

        edit_url = "#"
        if obj.status == "OG":
            edit_url = reverse("admin:cashbox-edit-out", args=[obj.id])
        elif obj.status == "CM":
            edit_url = reverse("admin:cashbox-edit-in", args=[obj.id])

        delete_url = reverse("admin:cashbox-delete", args=[obj.id])

        row["actions"] = format_html(
            '<div class="btn-group btn-group-sm">'
            '<a href="{}" class="btn btn-default"><i class="fa fa-pencil"></i></a>'
            '<a href="{}" class="btn btn-default"><i class="fa fa-trash"></i></a>'
            "</div>",
            edit_url,
            delete_url,
        )

        for key in row:
            if row[key] is None or row[key] == "":
                row[key] = '<span class="text-muted">(не задано)</span>'

    def get_initial_queryset(self, request=None):
        qs = super().get_initial_queryset().order_by("id")

        owner = self.request.POST.get("owner")
        date_from = self.request.POST.get("date_from")
        date_to = self.request.POST.get("date_to")
        status = self.request.POST.get("statuss")
        article = self.request.POST.get("article")
        bank_book_id = (
            self.request.POST.get("bank_book_id")
            or self.request.GET.get("bank_book_id")
            or self.request.POST.get("extra_data[bank_book_id]")
        )

        if bank_book_id:
            qs = qs.filter(bank_book_id=bank_book_id, status="CM")

        if owner:
            qs = qs.filter(owner=owner)

        if date_to:
            qs = qs.filter(date__lte=date_to)

        if date_from:
            qs = qs.filter(date__gte=date_from)

        if status:
            if status == "is_catch":
                qs = qs.filter(is_catch=True)
            elif status == "not_is_catch":
                qs = qs.filter(is_catch=False)

        if article:
            qs = qs.filter(article=article)

        return qs


class CashBoxDetail(RedirectMixin, PermissionRequiredMixin, DetailView):
    model = CashBox
    template_name = "cashbox_detail.html"
    context_object_name = "cashbox"
    permission_required = "role.has_cashbox"


class CashBoxDeleteView(RedirectMixin, PermissionRequiredMixin, DeleteView):
    model = "CashBox"
    permission_required = "role.has_cashbox"

    def get(self, request, pk):
        cashbox = get_object_or_404(CashBox, pk=pk)
        cashbox.delete()
        return redirect("admin:cashbox-list")


def cashbox_xlsx(request):
    cashbox = get_object_or_404(CashBox)

    wb = Workbook()
    ws = wb.active
    ws.title = "Платеж"

    data = [
        ("Платеж", f"#{cashbox.number}"),
        ("Дата", cashbox.date.strftime("%d.%m.%Y") if cashbox.date else ""),
        ("Владелец квартиры", str(cashbox.owner.fullname) if cashbox.owner else ""),
        ("Лицевой счет", str(cashbox.bank_book.number) if cashbox.bank_book else ""),
        ("Приход/расход", cashbox.get_status_display()),
        ("Статус", "Проведен" if cashbox.is_catch else "Не проведен"),
        ("Статья", str(cashbox.article.title) if cashbox.article else ""),
        ("Квитанция", str(cashbox.receipt.number) if cashbox.receipt else ""),
        ("Сумма", cashbox.sum),
        ("Валюта", "UAH"),
        ("Комментарий", cashbox.comment if cashbox.comment else ""),
        ("Менеджер", cashbox.manager.fullname),
    ]

    for row_num, (label, value) in enumerate(data, start=1):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=value)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    now = timezone.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename="cashbox_{now}.xlsx"'
    wb.save(response)

    return response


def cashbox_detail_xlsx(request, pk):
    wb = Workbook()
    ws = wb.active
    ws.title = "XLSX"
    columns = [
        "#",
        "Дата",
        "Приход/расход",
        "Статус",
        "Статья",
        "Квитанция",
        "Сумма",
        "Валюта",
        "Владелец квартиры",
        "Лицевой счет",
    ]
    ws.append(columns)
    qs = CashBox.objects.all()
    for cashbox in qs:
        row = [
            cashbox.number,
            cashbox.date.strftime("%d.%m.%Y"),
            cashbox.get_status_display(),
            "Проведен" if cashbox.is_catch else "Не проведен",
            cashbox.article.title if cashbox.article else "",
            cashbox.receipt.number if cashbox.receipt else "",
            cashbox.sum,
            "UAH",
            cashbox.owner.fullname if cashbox.owner else "",
            cashbox.bank_book.number if cashbox.bank_book else "",
        ]
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    now = timezone.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename="cashbox_{now}.xlsx"'
    wb.save(response)
    return response
